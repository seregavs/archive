from openpyxl.styles import NamedStyle, Font, Border, Side, Color, PatternFill, Alignment

bd = Side(style='thin', color="000000") # тонкий
colheader_regular = NamedStyle(name="colheader_regular",
                # font = Font(bold=True),
                border = Border(left=bd, top=bd, right=bd, bottom=bd),
                alignment = Alignment(horizontal = 'center'),                
                fill = PatternFill(patternType='solid', fgColor=Color(rgb='FF9AD4FF'))
                )
colheader_total = NamedStyle(name="colheader_total",
                # font = Font(bold=True),
                border = Border(left=bd, top=bd, right=bd, bottom=bd),
                alignment = Alignment(horizontal = 'center'),
                fill = PatternFill(patternType='solid', fgColor=Color(rgb='FF7BAACC'))
                )
rowheader_regular = NamedStyle(name="rowheader_regular",
                # font = Font(bold=True),
                border = Border(left=bd, top=bd, right=bd, bottom=bd),
                alignment = Alignment(horizontal = 'left'),
                fill = PatternFill(patternType='solid', fgColor=Color(rgb='FF9AD4FF'))
                )                
rowheader_total = NamedStyle(name="rowheader_total",
                # font = Font(bold=True),
                border = Border(left=bd, top=bd, right=bd, bottom=bd),
                alignment = Alignment(horizontal = 'left'),
                fill = PatternFill(patternType='solid', fgColor=Color(rgb='FF7BAACC'))
                )  
cellvalue_totlr = NamedStyle(name="cellvalue_totlr",
                font = Font(bold=True),
                border = Border(left=bd, top=bd, right=bd, bottom=bd),
                alignment = Alignment(horizontal = 'right'),
                fill = PatternFill(patternType='solid', fgColor=Color(rgb='FFE8E24A'))
                )
cellvalue_totlc = NamedStyle(name="cellvalue_totlc",
                font = Font(bold=True),
                border = Border(left=bd, top=bd, right=bd, bottom=bd),
                alignment = Alignment(horizontal = 'right'),
                fill = PatternFill(patternType='solid', fgColor=Color(rgb='FFEAED96'))
                )                
cellvalue_regular = NamedStyle(name="cellvalue_regular",
                # font = Font(bold=True),
                border = Border(left=bd, top=bd, right=bd, bottom=bd),
                alignment = Alignment(horizontal = 'right'),
                fill = PatternFill(patternType='solid', fgColor=Color(rgb='FFFFFFFF'))
                )