from openpyxl import comments
import pandas as pd
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.comments import Comment
# from openpyxl.utils import get_column_letter
# from openpyxl.styles import NamedStyle, Font, Border, Side, Color, PatternFill, Alignment
import excelstyles01 as exs01

os.system('cls' if os.name == 'nt' else 'clear')

src =  ['1;1;2020;1;10;30',
        '1;1;2020;2;11;31',
        '1;1;2021;1;12;32',
        '1;1;2021;2;13;33',
        '1;2;2020;1;14;34',
        '1;2;2020;2;15;35',
        '1;2;2021;1;16;36',
        '1;2;2021;2;17;37',
        '2;1;2020;1;18;38',
        '2;1;2020;2;19;39',
        '2;1;2021;1;20;40',
        '2;1;2021;2;21;41',
        '2;2;2020;1;22;42',
        '2;2;2020;2;23;43',
        '2;2;2021;1;24;44',
        '2;2;2021;2;25;45',
        ]
char_count  = 4
column_count = src[0].count(";")+1

src = list(map(lambda s: s.split(';'), src))
src = list(map(lambda row: list(map(lambda e, i: float(e) if i >= char_count else e, row, range(len(row)))), src))
# src = list(zip(*src))
# print(src)
df = pd.DataFrame.from_records(src, columns = list(map(lambda i: chr(65+i), range(column_count))))                
table = pd.pivot_table(df, values='E', index=['A', 'B'],
                    columns=['C', 'D'], aggfunc=np.sum, margins = False)

# print(table)
print('\ngroupby_index=\n',list(table.groupby(level=0, axis='index')))
table = pd.concat([d.append(d.sum(axis='index').rename((k, k + ' Total')))  for k, d in table.groupby(level=0, axis='index')
                  ], axis='index').append(table.sum(axis='index').rename(('Grand', 'Grand')))
print('\nFinal result after row joins\n', table)
print('\ngroupby_columns=\n',list(table.groupby(level=0, axis='columns')))
print('$$$$$$$$$$$ PLAY WITH GROUP BY COLS %%%%%%%%%%')
dflist = []
for k, d in table.groupby(level=0, axis='columns'):
        # print(k)
        # print(d)
        # print(d.sum(axis='columns').rename((k,'Subtotal-'+k)))
        dflist.append(d.join([d.sum(axis='columns').rename((k,'Subtotal-'+k))], rsuffix = 'T'))
# https://pandas.pydata.org/pandas-docs/stable/user_guide/merging.html      
table = pd.concat(dflist,axis='columns', join='inner').join(table.sum(axis='columns').rename(('Grand','Grand')))
print('Final result after row and column join\n',table)

# print('***** TOTAL on COLUMNS **********')        
# print(table.sum(axis='columns').rename(('Grand','Grand!!')))
print('********** Excel part ***********')

# Prints by cell. Ready to save in Excel
# for a in table.columns.to_flat_index():
#         print('column: ', a, a[0], a[1])
# for ind, data in table.iterrows():
#         print(f'index=: {ind}')
#         print(f'data=: {list(data)}')


replayoutList = []
replayoutList.append({'name' : 'Report1', 'rTop' : 5, 'cTop': 4, 'repTitleCell':'D2' })

wb = Workbook()
dest_filename = 'BroadcastExcel.xlsx'
ws2 = wb.create_sheet(title=replayoutList[0]['name'])
ws2.sheet_properties.tabColor = "1072BA"

comment = Comment(r'Это комментарий к названию отчета','I028159')
cell = ws2[replayoutList[0]['repTitleCell']]
cell.value = replayoutList[0]['name']
cell.comment = comment

wb.add_named_style(exs01.colheader_regular)
wb.add_named_style(exs01.colheader_total)
wb.add_named_style(exs01.rowheader_regular)
wb.add_named_style(exs01.rowheader_total)
wb.add_named_style(exs01.cellvalue_regular)
wb.add_named_style(exs01.cellvalue_totlr)
wb.add_named_style(exs01.cellvalue_totlc)


# >>> comment = ws["A1"].comment
# >>> comment = Comment('This is the comment text', 'Comment Author')
# >>> comment.text



i = int(replayoutList[0]['cTop']) + 2
for a in table.columns.to_flat_index():
        cell = ws2.cell(column = i, row = int(replayoutList[0]['rTop']), value=str(a[0]))
        cell.style = exs01.colheader_total if ( 'Subtotal' in str(a[0])) | ( 'Grand' in str(a[0])) else exs01.colheader_regular
        cell = ws2.cell(column = i, row = int(replayoutList[0]['rTop'])+1, value=str(a[1]))
        cell.style = exs01.colheader_total if ( 'Subtotal' in str(a[1])) | ( 'Grand' in str(a[1])) else exs01.colheader_regular
        i += 1

j = int(replayoutList[0]['rTop']) + 2        
for ind, data in table.iterrows():
        i = int(replayoutList[0]['cTop']) + 2
        cell = ws2.cell(column = i-2, row = j, value=str(ind[0]))
        cell.style = exs01.rowheader_total if ( 'Total' in str(ind[0])) | ( 'Grand' in str(ind[0])) else exs01.rowheader_regular 
        cell = ws2.cell(column = i-1, row = j, value=str(ind[1]))
        cell.style = exs01.rowheader_total if ( 'Total' in str(ind[1])) | ( 'Grand' in str(ind[1])) else exs01.rowheader_regular   
        for cv in data:
                cell = ws2.cell(column = i, row = j, value=str(cv))                
                colval = str(ws2.cell(row = int(replayoutList[0]['rTop'])+1, column = i).value)
                colval = 'COLT' if ('Subtotal' in colval ) | ('Grand' in colval ) else 'COLR'
                if colval == 'COLT':             
                        cell.style = exs01.cellvalue_totlr if ( 'Total' in str(ind[1])) | ( 'Grand' in str(ind[1])) else exs01.cellvalue_totlc 
                else:
                        cell.style = exs01.cellvalue_totlr if ( 'Total' in str(ind[1])) | ( 'Grand' in str(ind[1])) else exs01.cellvalue_regular
                i += 1     
        j += 1
for i in range(int(replayoutList[0]['cTop']),int(replayoutList[0]['cTop'])+2,1):
        for j in range (int(replayoutList[0]['rTop']),int(replayoutList[0]['rTop'])+2,1):
                ws2.cell(column = i, row = j).style = exs01.rowheader_regular                

# ws2.auto_filter.ref = "C2:K9"      

# To do:
# autofilter automatic calculation
# autowidth of columns ( https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size, DimensionHolder)
# insert image from file`
# merge cells (column header)
# freeze top n (https://stackoverflow.com/questions/25588918/how-to-freeze-entire-header-row-in-openpyxl)
### move excel styles to separate file
# declare lefttop as constants
# declare Grand and subtotal as constants
# vertical and horizontal grouping
# cumulative values (runnint totals)
### “tabColor”
# comments add 
# chart (?)
# test on large array 100000 x 12
# read from config file (see below). No JSON!!!


# https://openpyxl.readthedocs.io/en/stable/index.html

# ws2['F5'] = 3.14
# https://tproger.ru/translations/5-ways-of-using-underscore-in-python/
# ws3 = wb.create_sheet(title="Data")
# for row in range(10, 20):
#      for col in range(27, 54):
#          _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
# print(ws3['AA10'].value)
wb.save(filename = dest_filename)

# from configparser import ConfigParser
# config = ConfigParser()

# config.read('config.ini')
# config.add_section('main')
# config.set('main', 'key1', 'value1')
# config.set('main', 'key2', 'value2')
# config.set('main', 'key3', 'value3')

# with open('config.ini', 'w') as f:
# config.write(f)
# The file format is very simple with sections marked out in square brackets:

# [main]
# key1 = value1
# key2 = value2
# key3 = value3


# print(type(table))
# https://www.reddit.com/r/learnpython/comments/iwciyk/how_to_display_subtotals_for_columns_using_pandas/
# *********************************************** ANOTHER EXAMPlE OF SUBTOTAL in ROWs ONLY ******************************
# https://stackoverflow.com/questions/41383302/pivot-table-subtotals-in-pandas
# ******************** SOLUTION to AGGR of 2 and more key figures at once ****************************
# https://stackoverflow.com/questions/44165629/pandas-pivot-table-for-multiple-columns-at-once